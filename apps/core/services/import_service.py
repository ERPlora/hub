"""
Import service for ERPlora Hub.

Provides CSV and Excel import functionality for file uploads.
All modules should use this service for consistent data import.
"""

import csv
import io
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

from django.core.files.uploadedfile import UploadedFile
from django.utils.translation import gettext as _


@dataclass
class ImportResult:
    """Result of an import operation."""
    created: int = 0
    skipped: int = 0
    errors: List[Tuple[int, str]] = field(default_factory=list)


def parse_import_file(
    file: UploadedFile,
    expected_headers: Optional[List[str]] = None,
) -> List[dict]:
    """
    Parse an uploaded CSV or Excel file into a list of dicts.

    Detects format by file extension. Returns rows as dicts keyed by
    header names (first row of the file).

    Args:
        file: Django UploadedFile object
        expected_headers: Optional list of required headers to validate

    Returns:
        List of dicts, one per data row

    Raises:
        ValueError: If file format is unsupported or headers are missing

    Example:
        >>> from apps.core.services.import_service import parse_import_file
        >>> rows = parse_import_file(request.FILES['file'])
        >>> for row in rows:
        ...     print(row['Name'], row['Email'])
    """
    filename = file.name.lower()

    if filename.endswith('.csv'):
        rows = _parse_csv(file)
    elif filename.endswith(('.xlsx', '.xls')):
        rows = _parse_excel(file)
    else:
        raise ValueError(_("Unsupported file format. Please upload a CSV or Excel file."))

    if expected_headers and rows:
        actual = set(rows[0].keys())
        missing = set(expected_headers) - actual
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

    return rows


def _parse_csv(file: UploadedFile) -> List[dict]:
    """Parse a CSV file into list of dicts."""
    content = file.read()

    # Try UTF-8 with BOM, then UTF-8, then latin-1
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    else:
        raise ValueError(_("Could not decode file. Please ensure it is UTF-8 encoded."))

    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _parse_excel(file: UploadedFile) -> List[dict]:
    """Parse an Excel file into list of dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

    wb = load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active

    rows = []
    headers = None

    for row in sheet.iter_rows(values_only=True):
        if headers is None:
            headers = [str(cell).strip() if cell is not None else '' for cell in row]
            continue

        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(cell).strip() if cell is not None else ''

        rows.append(row_dict)

    wb.close()
    return rows


__all__ = [
    'parse_import_file',
    'ImportResult',
]
