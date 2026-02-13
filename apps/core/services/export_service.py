"""
Export service for ERPlora Hub.

Provides CSV and Excel export functionality for querysets and data.
All modules should use this service for consistent data export.
"""

import csv
import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Callable, Union

from django.http import HttpResponse
from django.db.models import QuerySet


def export_to_csv(
    data: Union[QuerySet, List[Dict]],
    fields: List[str],
    filename: Optional[str] = None,
    headers: Optional[List[str]] = None,
    field_formatters: Optional[Dict[str, Callable]] = None
) -> HttpResponse:
    """
    Export data to CSV file.

    Args:
        data: QuerySet or list of dicts to export
        fields: List of field names to include
        filename: Output filename (defaults to 'export_YYYYMMDD.csv')
        headers: Custom headers (defaults to field names)
        field_formatters: Dict of field -> formatter function

    Returns:
        HttpResponse: CSV file download response

    Example:
        >>> from apps.core.services.export_service import export_to_csv
        >>> products = Product.objects.all()
        >>> return export_to_csv(
        ...     products,
        ...     fields=['name', 'price', 'stock'],
        ...     filename='products.csv',
        ...     headers=['Product Name', 'Price (€)', 'In Stock']
        ... )
    """
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    if headers is None:
        headers = fields

    if field_formatters is None:
        field_formatters = {}

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Add BOM for Excel UTF-8 compatibility
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(headers)

    for item in data:
        row = []
        for field in fields:
            # Get value from dict or object
            if isinstance(item, dict):
                value = item.get(field, '')
            else:
                value = getattr(item, field, '')

            # Handle callable (e.g., methods)
            if callable(value):
                value = value()

            # Apply custom formatter if exists
            if field in field_formatters:
                value = field_formatters[field](value)
            else:
                value = _format_value(value)

            row.append(value)

        writer.writerow(row)

    return response


def export_to_excel(
    data: Union[QuerySet, List[Dict]],
    fields: List[str],
    filename: Optional[str] = None,
    headers: Optional[List[str]] = None,
    sheet_name: str = 'Data',
    field_formatters: Optional[Dict[str, Callable]] = None
) -> HttpResponse:
    """
    Export data to Excel file (.xlsx).

    Requires openpyxl to be installed.

    Args:
        data: QuerySet or list of dicts to export
        fields: List of field names to include
        filename: Output filename (defaults to 'export_YYYYMMDD.xlsx')
        headers: Custom headers (defaults to field names)
        sheet_name: Name of the Excel sheet
        field_formatters: Dict of field -> formatter function

    Returns:
        HttpResponse: Excel file download response

    Example:
        >>> return export_to_excel(
        ...     products,
        ...     fields=['name', 'price', 'stock'],
        ...     filename='products.xlsx'
        ... )
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    if headers is None:
        headers = fields

    if field_formatters is None:
        field_formatters = {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    row_count = 0
    for row_num, item in enumerate(data, 2):
        row_count += 1
        for col_num, field in enumerate(fields, 1):
            # Get value from dict or object
            if isinstance(item, dict):
                value = item.get(field, '')
            else:
                value = getattr(item, field, '')

            # Handle callable
            if callable(value):
                value = value()

            # Apply custom formatter if exists
            if field in field_formatters:
                value = field_formatters[field](value)
            else:
                value = _format_value_for_excel(value)

            sheet.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_num, _ in enumerate(headers, 1):
        max_length = len(str(headers[col_num - 1]))
        for row_num in range(2, row_count + 2):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        sheet.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


def _format_value(value: Any) -> str:
    """Format a value for CSV output."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _format_value_for_excel(value: Any) -> Any:
    """Format a value for Excel output (preserves types)."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, Decimal):
        return float(value)
    return value


def generate_csv_string(
    data: Union[QuerySet, List[Dict]],
    fields: List[str],
    headers: Optional[List[str]] = None
) -> str:
    """
    Generate CSV as a string (useful for storing or processing).

    Args:
        data: QuerySet or list of dicts
        fields: List of field names
        headers: Custom headers

    Returns:
        str: CSV content as string
    """
    if headers is None:
        headers = fields

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for item in data:
        row = []
        for field in fields:
            if isinstance(item, dict):
                value = item.get(field, '')
            else:
                value = getattr(item, field, '')

            if callable(value):
                value = value()

            row.append(_format_value(value))

        writer.writerow(row)

    return output.getvalue()


__all__ = [
    'export_to_csv',
    'export_to_excel',
    'generate_csv_string',
]
